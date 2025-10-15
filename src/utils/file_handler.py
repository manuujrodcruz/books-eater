"""File handling utilities for reading and writing data."""

import os
from typing import List, Optional
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from ..models.book import Book


class FileHandler:
    """
    Handles file I/O operations.
    """
    
    @staticmethod
    def load_books_from_file(filename: str) -> Optional[List[Book]]:
        """
        Load book list from a text file.
        
        Format: Título | Autor | Año
        
        Args:
            filename: Path to the books file
            
        Returns:
            List of Book objects or None if file not found
        """
        try:
            if not os.path.exists(filename):
                return None
            
            books = []
            with open(filename, 'r', encoding='utf-8') as f:
                for idx, line in enumerate(f, 1):
                    line = line.strip()
                    
                    # Skip empty lines and comments
                    if not line or line.startswith('#'):
                        continue
                    
                    book = Book.create_from_text(idx, line)
                    if book:
                        books.append(book)
            
            return books if books else None
            
        except Exception as e:
            print(f"Error leyendo {filename}: {e}")
            return None
    
    @staticmethod
    def save_to_excel(books: List[Book], filename: str) -> bool:
        """
        Save books to Excel file with formatting.
        
        Args:
            books: List of Book objects
            filename: Output filename
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Convert books to dictionaries
            data = [book.to_dict() for book in books]
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Write to Excel with formatting
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Audiolibros Dominicanos')
                
                worksheet = writer.sheets['Audiolibros Dominicanos']
                
                # Set column widths
                column_widths = {
                    'A': 10,  # Número
                    'B': 40,  # Título Libro
                    'C': 30,  # Autor
                    'D': 10,  # Año
                    'E': 60,  # URL YouTube
                    'F': 15,  # Duración
                    'G': 25,  # Tipo Contenido
                    'H': 15,  # Disponibilidad
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Format header row
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                header_font = Font(bold=True, color='FFFFFF')
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Format data rows
                for row in range(2, len(books) + 2):
                    # Align cells
                    worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='center')  # Número
                    worksheet.cell(row=row, column=4).alignment = Alignment(horizontal='center')  # Año
                    worksheet.cell(row=row, column=6).alignment = Alignment(horizontal='center')  # Duración
                    worksheet.cell(row=row, column=8).alignment = Alignment(horizontal='center')  # Disponibilidad
                    
                    # Color code availability
                    disponibilidad = worksheet.cell(row=row, column=8).value
                    if disponibilidad == "ENCONTRADO":
                        worksheet.cell(row=row, column=8).fill = PatternFill(
                            start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'
                        )
                        worksheet.cell(row=row, column=8).font = Font(color='006100')
                    elif disponibilidad == "PARCIAL":
                        worksheet.cell(row=row, column=8).fill = PatternFill(
                            start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'
                        )
                        worksheet.cell(row=row, column=8).font = Font(color='9C5700')
                    else:  # NO ENCONTRADO
                        worksheet.cell(row=row, column=8).fill = PatternFill(
                            start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
                        )
                        worksheet.cell(row=row, column=8).font = Font(color='9C0006')
            
            print(f"Excel guardado exitosamente: {filename}")
            return True
            
        except Exception as e:
            print(f"Error guardando Excel: {e}")
            return False
    
    @staticmethod
    def save_to_csv(books: List[Book], filename: str) -> bool:
        """
        Save books to CSV file.
        
        Args:
            books: List of Book objects
            filename: Output filename
            
        Returns:
            True if successful, False otherwise
        """
        try:
            data = [book.to_dict() for book in books]
            df = pd.DataFrame(data)
            df.to_csv(filename, index=False, encoding='utf-8')
            
            print(f"CSV guardado exitosamente: {filename}")
            return True
            
        except Exception as e:
            print(f"Error guardando CSV: {e}")
            return False
